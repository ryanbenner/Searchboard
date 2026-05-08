from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from jobscraper.job import Job


_HEADERS = ["Score", "Company", "Title", "Location", "Remote",
            "Salary min", "Salary max", "Source", "URL", "Rationale"]


def _row(j: Job) -> list:
    return [
        j.score, j.company, j.title, j.location,
        "Y" if j.remote else "",
        j.salary_min, j.salary_max, j.source, j.url, j.rationale or "",
    ]


def _add_sheet(wb: Workbook, name: str, jobs: list[Job]) -> None:
    ws = wb.create_sheet(name)
    ws.append(_HEADERS)
    for j in sorted(jobs, key=lambda x: (x.score is None, -(x.score or 0))):
        ws.append(_row(j))
    if ws.max_row >= 2:
        rule = ColorScaleRule(
            start_type="num", start_value=0,  start_color="F8696B",
            mid_type="num",   mid_value=60,   mid_color="FFEB84",
            end_type="num",   end_value=100,  end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"A2:A{ws.max_row}", rule)
    for i, _ in enumerate(_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def write_xlsx(path: str | Path, *,
               new_today: list[Job], still_open: list[Job],
               all_ranked: list[Job]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, "New today", new_today)
    _add_sheet(wb, "Still open", still_open)
    _add_sheet(wb, "All ranked", all_ranked)
    wb.save(path)


def render_markdown(jobs: list[Job], *, top_n: int = 15) -> str:
    ranked = sorted(jobs, key=lambda x: (x.score is None, -(x.score or 0)))[:top_n]
    if not ranked:
        return "_No new jobs today._"
    lines = ["# JobScraper — top jobs today", ""]
    for j in ranked:
        salary = ""
        if j.salary_min and j.salary_max:
            salary = f" — ${j.salary_min:,}–${j.salary_max:,}"
        score = j.score if j.score is not None else "?"
        lines.append(
            f"- **[{j.title}]({j.url})** at *{j.company}* "
            f"({j.location}){salary} — score **{score}**. "
            f"{j.rationale or ''}"
        )
    return "\n".join(lines)
