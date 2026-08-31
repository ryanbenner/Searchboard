"""
End-to-end pipeline test using only canned fixtures + mocks. Asserts the
shape of the outputs (xlsx tabs, sqlite rows, email body) without hitting
network or Anthropic.
"""
import json
import shutil
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
import respx, httpx
from openpyxl import load_workbook
from searchboard import cli


FIX = Path(__file__).parent / "fixtures"


@respx.mock
def test_full_pipeline(tmp_path, monkeypatch):
    shutil.copy(FIX / "profile_min.yml", tmp_path / "profile.yml")
    shutil.copy(FIX / "companies_min.yml", tmp_path / "companies.yml")
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()

    respx.get("https://boards-api.greenhouse.io/v1/boards/anthropic/jobs").mock(
        return_value=httpx.Response(200, json=json.loads((FIX/"greenhouse_anthropic.json").read_text())))
    respx.get("https://api.lever.co/v0/postings/notion").mock(
        return_value=httpx.Response(200, json=json.loads((FIX/"lever_notion.json").read_text())))
    respx.get("https://remoteok.com/api").mock(
        return_value=httpx.Response(200, json=json.loads((FIX/"remoteok.json").read_text())))
    respx.get("https://remotive.com/api/remote-jobs").mock(
        return_value=httpx.Response(200, json=json.loads((FIX/"remotive.json").read_text())))
    respx.get(host="weworkremotely.com").mock(return_value=httpx.Response(404))
    respx.get(host="api.ashbyhq.com").mock(return_value=httpx.Response(404))
    respx.get("https://hn.algolia.com/api/v1/search").mock(return_value=httpx.Response(200, json={"hits": []}))

    monkeypatch.setattr(cli, "verify_links", lambda jobs, **kw: jobs)

    def fake_rank(jobs, profile, client=None):
        for j in jobs:
            j.score = 75
            j.rationale = "ok fit"
        return jobs
    monkeypatch.setattr(cli, "rank_jobs", fake_rank)

    sent = {}
    monkeypatch.setattr(cli, "send_digest", lambda **kw: sent.update(kw))

    for k in ("ANTHROPIC_API_KEY", "SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "EMAIL_TO"):
        monkeypatch.setenv(k, "x")
    monkeypatch.setenv("SMTP_PORT", "587")

    cli.main(["run"])

    xlsx = tmp_path / "data" / "latest.xlsx"
    assert xlsx.exists()
    wb = load_workbook(xlsx)
    assert wb.sheetnames == ["New today", "Still open", "All ranked"]
    assert wb["All ranked"].max_row >= 2

    sqlite = tmp_path / "data" / "seen.sqlite"
    assert sqlite.exists()

    assert sent["to"] == "x"
    assert "ok fit" in sent["markdown_body"]
