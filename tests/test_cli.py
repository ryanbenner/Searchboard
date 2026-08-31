import shutil
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock
from openpyxl import load_workbook
from searchboard import cli
from searchboard.job import Job


def _j(jid, title="Junior Software Engineer", score=None):
    return Job(
        id=jid, source=jid.split(":")[0], company="ACo", title=title,
        location="Remote", remote=True, salary_min=80000, salary_max=120000,
        url="https://x", posted_at=None, seen_at=date.today(),
        description_text="vue/node", score=score, rationale=None,
    )


def test_cli_run_end_to_end(tmp_path, monkeypatch):
    shutil.copy(Path(__file__).parent / "fixtures" / "profile_min.yml", tmp_path / "profile.yml")
    shutil.copy(Path(__file__).parent / "fixtures" / "companies_min.yml", tmp_path / "companies.yml")
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()

    fake_src = MagicMock()
    fake_src.fetch.return_value = [_j("greenhouse:anthropic:1")]
    monkeypatch.setattr(cli, "build_sources",
                        lambda profile, companies: [fake_src])
    monkeypatch.setattr(cli, "verify_links", lambda jobs, **kw: jobs)

    def fake_rank(jobs, profile, client=None):
        for j in jobs: j.score = 90; j.rationale = "great match"
        return jobs
    monkeypatch.setattr(cli, "rank_jobs", fake_rank)

    monkeypatch.setenv("SMTP_HOST", "smtp.example.com")
    monkeypatch.setenv("SMTP_PORT", "587")
    monkeypatch.setenv("SMTP_USER", "u@example.com")
    monkeypatch.setenv("SMTP_PASS", "p")
    monkeypatch.setenv("EMAIL_TO", "u@example.com")
    monkeypatch.setenv("ANTHROPIC_API_KEY", "fake")

    sent = {}
    def fake_send(**kw): sent.update(kw)
    monkeypatch.setattr(cli, "send_digest", fake_send)

    cli.main(["run"])

    xlsx = tmp_path / "data" / "latest.xlsx"
    assert xlsx.exists()
    wb = load_workbook(xlsx)
    assert wb["New today"].cell(2, 1).value == 90
    assert sent["to"] == "u@example.com"
    assert "great match" in sent["markdown_body"]

    # Second run with the same job should NOT re-email — sent queue drained.
    sent.clear()
    cli.main(["run"])
    assert sent == {}, "second run must not email already-sent jobs"


def test_cli_skips_email_when_no_unsent_jobs(tmp_path, monkeypatch):
    shutil.copy(Path(__file__).parent / "fixtures" / "profile_min.yml", tmp_path / "profile.yml")
    shutil.copy(Path(__file__).parent / "fixtures" / "companies_min.yml", tmp_path / "companies.yml")
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir()

    fake_src = MagicMock()
    fake_src.fetch.return_value = [_j("greenhouse:anthropic:1")]
    monkeypatch.setattr(cli, "build_sources",
                        lambda profile, companies: [fake_src])
    monkeypatch.setattr(cli, "verify_links", lambda jobs, **kw: jobs)
    # Score below floor → not eligible
    monkeypatch.setattr(cli, "rank_jobs",
                        lambda jobs, profile, client=None: [
                            setattr(j, "score", 30) or j for j in jobs
                        ])

    for k in ("SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "EMAIL_TO", "ANTHROPIC_API_KEY"):
        monkeypatch.setenv(k, "x")
    monkeypatch.setenv("SMTP_PORT", "587")

    called = []
    monkeypatch.setattr(cli, "send_digest", lambda **kw: called.append(kw))

    cli.main(["run"])
    assert called == [], "should not send email when no eligible jobs"


from searchboard.cli import build_parser


def test_run_flag_defaults():
    args = build_parser().parse_args(["run"])
    assert args.no_email is False
    assert args.profile == "profile.yml"
    assert args.data_dir == "data"


def test_run_flags_parse():
    args = build_parser().parse_args(
        ["run", "--no-email", "--profile", "/d/profile.yml", "--data-dir", "/d"])
    assert args.no_email is True
    assert args.profile == "/d/profile.yml"
    assert args.data_dir == "/d"


def _spy_run(tmp_path, monkeypatch, rank_calls):
    shutil.copy(Path(__file__).parent / "fixtures" / "profile_min.yml", tmp_path / "profile.yml")
    shutil.copy(Path(__file__).parent / "fixtures" / "companies_min.yml", tmp_path / "companies.yml")
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data").mkdir(exist_ok=True)

    fake_src = MagicMock()
    fake_src.fetch.return_value = [_j("greenhouse:anthropic:1"), _j("greenhouse:anthropic:2")]
    monkeypatch.setattr(cli, "build_sources", lambda profile, companies: [fake_src])
    monkeypatch.setattr(cli, "verify_links", lambda jobs, **kw: jobs)

    def fake_rank(jobs, profile, client=None):
        rank_calls.append([j.id for j in jobs])
        for j in jobs:
            j.score = 90
            j.rationale = "great match"
        return jobs
    monkeypatch.setattr(cli, "rank_jobs", fake_rank)


def test_second_run_skips_already_ranked_jobs(tmp_path, monkeypatch):
    rank_calls = []
    _spy_run(tmp_path, monkeypatch, rank_calls)
    cli.main(["run", "--no-email"])
    cli.main(["run", "--no-email"])
    assert rank_calls[0] == ["greenhouse:anthropic:1", "greenhouse:anthropic:2"]
    assert rank_calls[1] == []
    # reused scores still land in the store on the second run
    from searchboard.store import Store
    row = Store(tmp_path / "data" / "seen.sqlite")._conn.execute(
        "SELECT ranked_score, rationale FROM seen WHERE id='greenhouse:anthropic:1'").fetchone()
    assert row["ranked_score"] == 90
    assert row["rationale"] == "great match"


def test_profile_change_triggers_full_rerank(tmp_path, monkeypatch):
    rank_calls = []
    _spy_run(tmp_path, monkeypatch, rank_calls)
    cli.main(["run", "--no-email"])
    with open(tmp_path / "profile.yml", "a") as f:
        f.write("# tweaked\n")
    cli.main(["run", "--no-email"])
    assert rank_calls[1] == ["greenhouse:anthropic:1", "greenhouse:anthropic:2"]


def test_stale_postings_stay_out_of_digest_and_xlsx(tmp_path, monkeypatch):
    rank_calls = []
    _spy_run(tmp_path, monkeypatch, rank_calls)
    from datetime import timedelta
    stale = _j("greenhouse:anthropic:1")
    stale.posted_at = date.today() - timedelta(days=40)
    stale.url = "https://stale"
    fresh = _j("greenhouse:anthropic:2")
    fresh.posted_at = date.today() - timedelta(days=2)
    fresh.url = "https://fresh"
    cli.build_sources(None, None)[0].fetch.return_value = [stale, fresh]
    sent = {}
    monkeypatch.setattr(cli, "send_digest", lambda **kw: sent.update(kw))
    for k in ("SMTP_HOST", "SMTP_USER", "SMTP_PASS", "EMAIL_TO"):
        monkeypatch.setenv(k, "x")
    monkeypatch.setenv("SMTP_PORT", "587")
    cli.main(["run"])
    assert "https://fresh" in sent["markdown_body"]
    assert "https://stale" not in sent["markdown_body"]
    wb = load_workbook(tmp_path / "data" / "latest.xlsx")
    urls = [ws.cell(r, 9).value for ws in wb.worksheets for r in range(2, ws.max_row + 1)]
    assert "https://fresh" in urls
    assert "https://stale" not in urls
