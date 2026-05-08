from datetime import date
from openpyxl import load_workbook
from jobscraper.digest import write_xlsx, render_markdown
from jobscraper.job import Job


def _j(jid, title, score):
    return Job(
        id=jid, source="x", company="ACo", title=title,
        location="Remote (US)", remote=True, salary_min=80000, salary_max=120000,
        url=f"https://x/{jid}", posted_at=None, seen_at=date.today(),
        description_text="", score=score, rationale="ok fit",
    )


def test_xlsx_has_three_tabs(tmp_path):
    new = [_j("x:y:1", "Junior FS", 88)]
    still = [_j("x:y:2", "Frontend", 72)]
    all_ranked = new + still
    out = tmp_path / "x.xlsx"
    write_xlsx(out, new_today=new, still_open=still, all_ranked=all_ranked)
    wb = load_workbook(out)
    assert wb.sheetnames == ["New today", "Still open", "All ranked"]
    ws = wb["New today"]
    assert ws.cell(1, 1).value == "Score"
    assert ws.cell(2, 1).value == 88


def test_markdown_lists_top_n():
    new = [_j(f"x:y:{i}", f"Job {i}", 100 - i) for i in range(20)]
    md = render_markdown(new, top_n=15)
    assert md.count("\n- ") == 15
    assert "Job 0" in md and "Job 14" in md
    assert "Job 15" not in md
